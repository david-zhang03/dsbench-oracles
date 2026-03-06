import openpyxl
wb = openpyxl.load_workbook('/home/ddzhang/DSBench/colab_tasks/dsbench_da_00000020/environment/data/MO16-Round-2-Sec-4-Maximize-the-Benefit.xlsx', data_only=True)
print('Sheet names:', wb.sheetnames)
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f'\n=== {sn} === (rows={ws.max_row}, cols={ws.max_column})')
    for row in ws.iter_rows(min_row=1, max_row=min(60, ws.max_row), values_only=False):
        vals = [(c.coordinate, c.value) for c in row if c.value is not None]
        if vals:
            print(vals)
