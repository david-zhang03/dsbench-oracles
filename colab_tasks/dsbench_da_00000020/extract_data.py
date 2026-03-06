#!/usr/bin/env python3
import openpyxl
import json

wb = openpyxl.load_workbook('/home/ddzhang/DSBench/colab_tasks/dsbench_da_00000020/environment/data/MO16-Round-2-Sec-4-Maximize-the-Benefit.xlsx', data_only=True)

result = {}
for sn in wb.sheetnames:
    ws = wb[sn]
    sheet_data = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
        row_data = []
        for c in row:
            row_data.append({'coord': c.coordinate, 'value': c.value})
        sheet_data.append(row_data)
    result[sn] = sheet_data

with open('/home/ddzhang/DSBench/colab_tasks/dsbench_da_00000020/extracted_data.json', 'w') as f:
    json.dump(result, f, indent=2, default=str)

print("Done extracting data")
