#!/usr/bin/env python3
import openpyxl
import json

wb = openpyxl.load_workbook('/home/ddzhang/DSBench/colab_tasks/dsbench_da_00000025/environment/data/MO15-Tax-Data (1).xlsx', data_only=True)
result = {"sheets": {}}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    sheet_data = {"max_row": ws.max_row, "max_col": ws.max_column, "rows": []}
    for row in ws.iter_rows(min_row=1, max_row=min(200, ws.max_row), values_only=False):
        row_data = {}
        for c in row:
            if c.value is not None:
                row_data[c.coordinate] = str(c.value)
        if row_data:
            sheet_data["rows"].append(row_data)
    result["sheets"][sheet_name] = sheet_data

with open('/home/ddzhang/DSBench/colab_tasks/dsbench_da_00000025/excel_data.json', 'w') as f:
    json.dump(result, f, indent=2, default=str)
print("Done - wrote excel_data.json")
