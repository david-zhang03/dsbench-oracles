import openpyxl

wb = openpyxl.load_workbook('/home/ddzhang/DSBench/colab_tasks/dsbench_da_00000025/environment/data/MO15-Tax-Data (1).xlsx', data_only=True)
print("Sheet names:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    for row in ws.iter_rows(min_row=1, max_row=min(50, ws.max_row), values_only=False):
        vals = [(c.coordinate, c.value) for c in row if c.value is not None]
        if vals:
            print(vals)
