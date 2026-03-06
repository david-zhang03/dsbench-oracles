#!/usr/bin/env python3
"""Explore the Excel data file for the Purple City task."""
import pandas as pd
import openpyxl

DATA_PATH = '/home/ddzhang/DSBench/colab_tasks/dsbench_da_00000011/environment/data/MO14-Purple-City.xlsx'

wb = openpyxl.load_workbook(DATA_PATH, data_only=True)
print("Sheet names:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet_name}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    print(f"{'='*60}")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        vals = [(cell.coordinate, cell.value) for cell in row if cell.value is not None]
        if vals:
            print(vals)
