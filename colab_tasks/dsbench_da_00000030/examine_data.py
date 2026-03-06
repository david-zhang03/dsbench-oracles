import openpyxl

wb = openpyxl.load_workbook('/home/ddzhang/DSBench/colab_tasks/dsbench_da_00000030/environment/data/MO14-Time-is-Money-Data.xlsx', data_only=True)
print("Sheet names:", wb.sheetnames)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n=== Sheet: {sheet_name} ===')
    print(f'Dimensions: {ws.dimensions}')
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
        vals = [(c.coordinate, c.value) for c in row if c.value is not None]
        if vals:
            print(vals)
