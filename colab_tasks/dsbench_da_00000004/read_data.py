import openpyxl
wb = openpyxl.load_workbook('/home/ddzhang/DSBench/colab_tasks/dsbench_da_00000004/environment/data/MO17-Round-1-Sec-4-Go-With-The-Flow-data.xlsx', data_only=True)
print('Sheet names:', wb.sheetnames)
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f'\nSheet: {sn}, rows={ws.max_row}, cols={ws.max_column}')
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80), values_only=False):
        vals = [(c.column_letter + str(c.row), c.value) for c in row if c.value is not None]
        if vals:
            print(vals)
